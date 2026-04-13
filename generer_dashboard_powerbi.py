#!/usr/bin/env python3
"""
Génère un fichier Excel multi-onglets prêt pour import Power BI.
Chaque onglet correspond à une table/vue du Data Warehouse.
Power BI > Obtenir des données > Excel > sélectionner ce fichier.
"""

import sys
import json
import logging
from pathlib import Path
from datetime import datetime

sys.path.insert(0, str(Path(__file__).resolve().parent))

logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
logger = logging.getLogger(__name__)

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.chart import BarChart, PieChart, Reference
    from openpyxl.chart.label import DataLabelList
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Installation de openpyxl...")
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.chart import BarChart, PieChart, Reference
    from openpyxl.chart.label import DataLabelList
    from openpyxl.utils import get_column_letter

from src.mongodb.gestionnaire_mongo import GestionnaireMongo

OUTPUT_DIR = Path(__file__).resolve().parent / "docs"

# Palette de couleurs professionnelle
BLEU_FONCE = "1B3A5C"
BLEU_MOYEN = "2E75B6"
BLEU_CLAIR = "D6E4F0"
GRIS_CLAIR = "F2F2F2"
BLANC = "FFFFFF"
VERT = "548235"
ORANGE = "ED7D31"

HEADER_FONT = Font(name="Calibri", bold=True, color=BLANC, size=11)
HEADER_FILL = PatternFill(start_color=BLEU_FONCE, end_color=BLEU_FONCE, fill_type="solid")
TITRE_FONT = Font(name="Calibri", bold=True, color=BLEU_FONCE, size=16)
SOUS_TITRE_FONT = Font(name="Calibri", bold=True, color=BLEU_MOYEN, size=12)
KPI_FONT = Font(name="Calibri", bold=True, color=BLEU_FONCE, size=28)
KPI_LABEL_FONT = Font(name="Calibri", color="666666", size=10)
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)


def style_header_row(ws, row, max_col):
    """Applique le style d'en-tête sur une ligne."""
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER


def style_data_rows(ws, start_row, end_row, max_col):
    """Applique un style zébré aux lignes de données."""
    zebra_fill = PatternFill(start_color=GRIS_CLAIR, end_color=GRIS_CLAIR, fill_type="solid")
    for row in range(start_row, end_row + 1):
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center")
            if (row - start_row) % 2 == 1:
                cell.fill = zebra_fill


def auto_width(ws, max_col, min_width=12, max_width=35):
    """Ajuste la largeur des colonnes automatiquement."""
    for col in range(1, max_col + 1):
        letter = get_column_letter(col)
        max_len = min_width
        for row in ws.iter_rows(min_col=col, max_col=col):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, min(len(str(cell.value)) + 2, max_width))
        ws.column_dimensions[letter].width = max_len


def creer_onglet_dashboard(wb, stats, donnees_type, donnees_dept):
    """Crée l'onglet Dashboard principal avec KPIs et graphiques."""
    ws = wb.active
    ws.title = "Dashboard"
    ws.sheet_properties.tabColor = BLEU_FONCE

    # Titre
    ws.merge_cells("A1:H1")
    ws["A1"] = "TABLEAU DE BORD — Hébergements Touristiques"
    ws["A1"].font = TITRE_FONT
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 40

    ws.merge_cells("A2:H2")
    ws["A2"] = f"Données actualisées le {datetime.now().strftime('%d/%m/%Y')}"
    ws["A2"].font = Font(name="Calibri", italic=True, color="999999", size=10)
    ws["A2"].alignment = Alignment(horizontal="center")

    # --- KPIs ---
    ws.merge_cells("A4:H4")
    ws["A4"] = "INDICATEURS CLÉS"
    ws["A4"].font = SOUS_TITRE_FONT

    kpis = [
        ("Total hébergements", stats.get("total", 0)),
        ("Types différents", len(donnees_type)),
        ("Départements", len(donnees_dept)),
        ("Communes", stats.get("total_communes", "—")),
    ]

    for i, (label, valeur) in enumerate(kpis):
        col = (i * 2) + 1
        ws.merge_cells(start_row=6, start_column=col, end_row=6, end_column=col + 1)
        ws.merge_cells(start_row=7, start_column=col, end_row=7, end_column=col + 1)
        cell_val = ws.cell(row=6, column=col, value=valeur)
        cell_val.font = KPI_FONT
        cell_val.alignment = Alignment(horizontal="center")
        cell_label = ws.cell(row=7, column=col, value=label)
        cell_label.font = KPI_LABEL_FONT
        cell_label.alignment = Alignment(horizontal="center")

    # --- Graphique : Répartition par type ---
    ws.merge_cells("A10:D10")
    ws["A10"] = "RÉPARTITION PAR TYPE"
    ws["A10"].font = SOUS_TITRE_FONT

    ws.cell(row=11, column=1, value="Type d'hébergement")
    ws.cell(row=11, column=2, value="Nombre")
    ws.cell(row=11, column=3, value="Capacité moy.")
    style_header_row(ws, 11, 3)

    for i, item in enumerate(donnees_type):
        row = 12 + i
        ws.cell(row=row, column=1, value=item["_id"])
        ws.cell(row=row, column=2, value=item["count"])
        ws.cell(row=row, column=3, value=round(item.get("avg_capacite", 0) or 0, 1))

    end_type = 11 + len(donnees_type)
    style_data_rows(ws, 12, end_type, 3)

    # Graphique barres horizontales par type
    chart_type = BarChart()
    chart_type.type = "bar"
    chart_type.title = "Nombre d'hébergements par type"
    chart_type.y_axis.title = None
    chart_type.x_axis.title = None
    chart_type.style = 26
    chart_type.width = 22
    chart_type.height = 14

    data_ref = Reference(ws, min_col=2, min_row=11, max_row=end_type)
    cats_ref = Reference(ws, min_col=1, min_row=12, max_row=end_type)
    chart_type.add_data(data_ref, titles_from_data=True)
    chart_type.set_categories(cats_ref)
    chart_type.series[0].graphicalProperties.solidFill = BLEU_MOYEN
    chart_type.legend = None
    ws.add_chart(chart_type, "A" + str(end_type + 2))

    # --- Graphique : Répartition par département ---
    dept_start = end_type + 18
    ws.merge_cells(f"A{dept_start}:D{dept_start}")
    ws.cell(row=dept_start, column=1, value="RÉPARTITION PAR DÉPARTEMENT").font = SOUS_TITRE_FONT

    noms_dept = {
        "44": "Loire-Atlantique (44)",
        "49": "Maine-et-Loire (49)",
        "53": "Mayenne (53)",
        "72": "Sarthe (72)",
        "85": "Vendée (85)",
        "35": "Ille-et-Vilaine (35)",
    }

    header_row = dept_start + 1
    ws.cell(row=header_row, column=1, value="Département")
    ws.cell(row=header_row, column=2, value="Nombre")
    style_header_row(ws, header_row, 2)

    for i, item in enumerate(donnees_dept):
        row = header_row + 1 + i
        dept_code = str(item["_id"])
        ws.cell(row=row, column=1, value=noms_dept.get(dept_code, f"Dept. {dept_code}"))
        ws.cell(row=row, column=2, value=item["count"])

    end_dept = header_row + len(donnees_dept)
    style_data_rows(ws, header_row + 1, end_dept, 2)

    # Graphique camembert par département
    pie = PieChart()
    pie.title = "Répartition par département"
    pie.style = 26
    pie.width = 18
    pie.height = 14

    data_ref = Reference(ws, min_col=2, min_row=header_row, max_row=end_dept)
    cats_ref = Reference(ws, min_col=1, min_row=header_row + 1, max_row=end_dept)
    pie.add_data(data_ref, titles_from_data=True)
    pie.set_categories(cats_ref)
    pie.dataLabels = DataLabelList()
    pie.dataLabels.showPercent = True
    pie.dataLabels.showVal = True
    ws.add_chart(pie, f"A{end_dept + 2}")

    auto_width(ws, 8)


def creer_onglet_donnees(wb, donnees, mongo):
    """Crée l'onglet avec toutes les données détaillées."""
    ws = wb.create_sheet("Données détaillées")
    ws.sheet_properties.tabColor = BLEU_MOYEN

    colonnes = [
        "Nom", "Type", "Adresse", "Code postal", "Commune",
        "Département", "Latitude", "Longitude", "Capacité pers.",
        "Capacité ch.", "Classement", "Téléphone", "Email",
        "Site web", "Tarif min", "Tarif max", "Source"
    ]
    champs = [
        "nom", "type_hebergement", "adresse", "code_postal", "commune",
        "departement", "latitude", "longitude", "capacite_personnes",
        "capacite_chambres", "classement", "telephone", "email",
        "site_web", "tarif_min", "tarif_max", "source"
    ]

    for col, nom in enumerate(colonnes, 1):
        ws.cell(row=1, column=col, value=nom)
    style_header_row(ws, 1, len(colonnes))

    ws.auto_filter.ref = f"A1:{get_column_letter(len(colonnes))}1"

    for i, doc in enumerate(donnees):
        row = i + 2
        for col, champ in enumerate(champs, 1):
            ws.cell(row=row, column=col, value=doc.get(champ))

    style_data_rows(ws, 2, len(donnees) + 1, len(colonnes))
    auto_width(ws, len(colonnes))

    logger.info(f"  {len(donnees)} lignes exportées dans 'Données détaillées'")


def creer_onglet_par_type(wb, donnees_type):
    """Crée l'onglet d'analyse par type d'hébergement."""
    ws = wb.create_sheet("Analyse par type")
    ws.sheet_properties.tabColor = VERT

    colonnes = ["Type d'hébergement", "Nombre", "Capacité moyenne", "% du total"]
    for col, nom in enumerate(colonnes, 1):
        ws.cell(row=1, column=col, value=nom)
    style_header_row(ws, 1, len(colonnes))

    total = sum(item["count"] for item in donnees_type)
    for i, item in enumerate(donnees_type):
        row = i + 2
        ws.cell(row=row, column=1, value=item["_id"])
        ws.cell(row=row, column=2, value=item["count"])
        ws.cell(row=row, column=3, value=round(item.get("avg_capacite", 0) or 0, 1))
        pct = round(item["count"] / total * 100, 1) if total else 0
        cell = ws.cell(row=row, column=4, value=pct / 100)
        cell.number_format = '0.0%'

    style_data_rows(ws, 2, len(donnees_type) + 1, len(colonnes))
    auto_width(ws, len(colonnes))


def creer_onglet_par_departement(wb, donnees_dept):
    """Crée l'onglet d'analyse par département."""
    ws = wb.create_sheet("Analyse par département")
    ws.sheet_properties.tabColor = ORANGE

    noms_dept = {
        "44": "Loire-Atlantique", "49": "Maine-et-Loire",
        "53": "Mayenne", "72": "Sarthe", "85": "Vendée",
        "35": "Ille-et-Vilaine",
    }

    colonnes = ["Code", "Département", "Nombre", "% du total"]
    for col, nom in enumerate(colonnes, 1):
        ws.cell(row=1, column=col, value=nom)
    style_header_row(ws, 1, len(colonnes))

    total = sum(item["count"] for item in donnees_dept)
    for i, item in enumerate(donnees_dept):
        row = i + 2
        code = str(item["_id"])
        ws.cell(row=row, column=1, value=code)
        ws.cell(row=row, column=2, value=noms_dept.get(code, code))
        ws.cell(row=row, column=3, value=item["count"])
        pct = round(item["count"] / total * 100, 1) if total else 0
        cell = ws.cell(row=row, column=4, value=pct / 100)
        cell.number_format = '0.0%'

    style_data_rows(ws, 2, len(donnees_dept) + 1, len(colonnes))
    auto_width(ws, len(colonnes))


def creer_onglet_geo(wb, donnees):
    """Crée l'onglet cartographique pour Power BI Map."""
    ws = wb.create_sheet("Carte géographique")
    ws.sheet_properties.tabColor = "70AD47"

    colonnes = ["Nom", "Type", "Commune", "Département", "Latitude", "Longitude", "Capacité"]
    for col, nom in enumerate(colonnes, 1):
        ws.cell(row=1, column=col, value=nom)
    style_header_row(ws, 1, len(colonnes))

    row = 2
    for doc in donnees:
        lat = doc.get("latitude")
        lon = doc.get("longitude")
        if lat and lon:
            ws.cell(row=row, column=1, value=doc.get("nom"))
            ws.cell(row=row, column=2, value=doc.get("type_hebergement"))
            ws.cell(row=row, column=3, value=doc.get("commune"))
            ws.cell(row=row, column=4, value=doc.get("departement"))
            ws.cell(row=row, column=5, value=lat)
            ws.cell(row=row, column=6, value=lon)
            ws.cell(row=row, column=7, value=doc.get("capacite_personnes"))
            row += 1

    style_data_rows(ws, 2, row - 1, len(colonnes))
    auto_width(ws, len(colonnes))
    logger.info(f"  {row - 2} points géographiques exportés pour la carte")


def generer():
    """Génère le fichier Excel dashboard pour Power BI."""
    logger.info("=" * 60)
    logger.info("  GÉNÉRATION DU DASHBOARD POWER BI")
    logger.info("=" * 60)

    # Connexion MongoDB
    mongo = GestionnaireMongo()
    stats = mongo.obtenir_statistiques()
    donnees_type = stats.get("par_type", [])
    donnees_dept = stats.get("par_departement", [])

    # Récupérer toutes les données clean
    donnees = list(mongo._clean.find({}, {"_id": 0}).sort("commune", 1))

    # Compter les communes distinctes
    stats["total_communes"] = len(mongo._clean.distinct("commune"))

    logger.info(f"  {len(donnees)} hébergements récupérés")
    logger.info(f"  {len(donnees_type)} types, {len(donnees_dept)} départements")

    # Créer le workbook Excel
    wb = openpyxl.Workbook()

    creer_onglet_dashboard(wb, stats, donnees_type, donnees_dept)
    creer_onglet_donnees(wb, donnees, mongo)
    creer_onglet_par_type(wb, donnees_type)
    creer_onglet_par_departement(wb, donnees_dept)
    creer_onglet_geo(wb, donnees)

    # Sauvegarder
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    output_path = OUTPUT_DIR / f"Dashboard_PowerBI_{datetime.now().strftime('%Y%m%d')}.xlsx"
    wb.save(str(output_path))

    mongo.fermer()

    logger.info("")
    logger.info(f"  Fichier généré : {output_path}")
    logger.info("")
    logger.info("  IMPORT POWER BI :")
    logger.info("  1. Ouvrir Power BI Desktop")
    logger.info("  2. Obtenir des données > Excel")
    logger.info(f"  3. Sélectionner {output_path.name}")
    logger.info("  4. Cocher tous les onglets > Charger")
    logger.info("  5. L'onglet 'Carte géographique' permet de créer une carte Power BI")
    logger.info("")
    logger.info("  OU CONNEXION DIRECTE MARIADB :")
    logger.info("  1. Obtenir des données > MySQL")
    logger.info("  2. Serveur : localhost:3306")
    logger.info("  3. Base : dw_hebergements")
    logger.info("  4. Vue : vue_powerbi_hebergements (table denormalisee)")
    logger.info("=" * 60)


if __name__ == "__main__":
    generer()
